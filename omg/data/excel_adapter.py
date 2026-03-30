# FILE: omg/data/excel_adapter.py
# Excel Data Source Adapter — SEC 05 of Technical Specification
# ═══════════════════════════════════════════════════════════════════

from __future__ import annotations

from typing import Any, Dict, Generator, List, Optional, Union

import openpyxl
from loguru import logger

from omg.data.adapter import AbstractAdapter, ColumnMeta, ColumnType


class ExcelAdapter(AbstractAdapter):
    """Reads .xlsx and .xls files using openpyxl or xlrd."""

    def __init__(self, path: str, sheet_name: Optional[str] = None):
        self.path = path
        self.sheet_name = sheet_name
        self._wb: Optional[openpyxl.Workbook] = None
        self._ws: Any = None
        self._xlrd_wb: Any = None
        self._xlrd_ws: Any = None
        self._header: Optional[List[str]] = None
        self._row_count: int = 0
        self._columns: Optional[List[ColumnMeta]] = None
        self._is_xls = self.path.lower().endswith('.xls')
        self._data_cache: Optional[List[Dict[str, Any]]] = None

    def _open(self) -> None:
        """Open the workbook and parse the header."""
        if self._wb is not None or self._xlrd_wb is not None:
            return

        header_row: List[Any] = []
        if self._is_xls:
            try:
                import xlrd
                self._xlrd_wb = xlrd.open_workbook(self.path)
                if self.sheet_name:
                    self._xlrd_ws = self._xlrd_wb.sheet_by_name(self.sheet_name)
                else:
                    self._xlrd_ws = self._xlrd_wb.sheet_by_index(0)
                
                header_row = self._xlrd_ws.row_values(0)
                self._row_count = self._xlrd_ws.nrows - 1
            except ImportError:
                logger.error("xlrd not installed. Required for .xls files.")
                raise ValueError("Support for .xls requires 'xlrd' package. Please install it.")
            except Exception as e:
                logger.error(f"Failed to load .xls workbook {self.path}: {e}")
                raise ValueError(f"Could not open .xls file: {e}")
        else:
            try:
                self._wb = openpyxl.load_workbook(self.path, read_only=True, data_only=True)
                if self.sheet_name:
                    if self.sheet_name not in self._wb.sheetnames:
                        raise ValueError(f"Sheet '{self.sheet_name}' not found.")
                    self._ws = self._wb[self.sheet_name]
                else:
                    self._ws = self._wb.active
                
                # In read_only mode, active might be None if no sheet is active
                if self._ws is None:
                    self._ws = self._wb[self._wb.sheetnames[0]]

                # Build a list of valid data rows (skipping entirely empty rows)
                # Note: We load the sheet into memory here. For typical label tasks (1k-10k rows),
                # this is perfectly fine and much faster than re-iterating read_only sheets.
                all_raw_rows = list(self._ws.iter_rows(values_only=True))
                if not all_raw_rows:
                    raise ValueError("Excel sheet is empty")
                
                header_row = all_raw_rows[0]
                self._header = [str(c).strip() if c is not None else f"Column_{i}" for i, c in enumerate(header_row)]
                
                # Deduplicate headers
                seen: Dict[str, int] = {}
                final_header = []
                for h in self._header:
                    name = h
                    if name in seen:
                        seen[name] += 1
                        name = f"{name}_{seen[name]}"
                    else:
                        seen[name] = 0
                    final_header.append(name)
                self._header = final_header

                self._data_cache = []
                for row in all_raw_rows[1:]:
                    # Skip if row is entirely empty
                    if any(v is not None and str(v).strip() != "" for v in row):
                        values = [str(v) if v is not None else "" for v in row]
                        self._data_cache.append(dict(zip(self._header, values)))
                
                self._row_count = len(self._data_cache)
                return # Skip redundant header logic below
            except Exception as e:
                logger.error(f"Failed to load .xlsx workbook {self.path}: {e}")
                raise ValueError(f"Could not open .xlsx file: {e}")

        # Deduplicate and clean headers
        seen: Dict[str, int] = {}
        self._header = []
        for i, c in enumerate(header_row):
            name = str(c).strip() if c is not None else f"Column_{i}"
            if name in seen:
                seen[name] += 1
                name = f"{name}_{seen[name]}"
            else:
                seen[name] = 0
            self._header.append(name)

        logger.info(f"Excel opened: {self.path} — {len(self._header)} columns, {self._row_count} rows")

    def _infer_column_types(self, sample_rows: int = 20) -> None:
        """Infer column types from the first N rows."""
        if self._columns is not None:
            return

        self._open()
        assert self._header is not None

        samples: dict[str, list[str]] = {col: [] for col in self._header}
        limit = min(sample_rows, self._row_count)

        for i in range(limit):
            row = self.get_row(i)
            for col in self._header:
                val = row.get(col, "")
                if val:
                    samples[col].append(str(val))

        self._columns = []
        for col in self._header:
            col_type = self._infer_type(samples[col])
            self._columns.append(ColumnMeta(
                name=col,
                inferred_type=col_type,
                sample_values=samples[col][:5],
            ))

    @staticmethod
    def _infer_type(values: list[str]) -> ColumnType:
        """Infer column type from sample values."""
        if not values:
            return ColumnType.STRING
        int_count = 0
        float_count = 0
        for v in values:
            v = v.strip()
            try:
                int(v)
                int_count += 1
                continue
            except ValueError:
                pass
            try:
                float(v)
                float_count += 1
                continue
            except ValueError:
                pass
        total = len(values)
        if int_count == total:
            return ColumnType.INTEGER
        if (int_count + float_count) == total and float_count > 0:
            return ColumnType.FLOAT
        return ColumnType.STRING

    # ── AbstractAdapter Interface ──

    def get_columns(self) -> List[ColumnMeta]:
        self._infer_column_types()
        assert self._columns is not None
        return self._columns

    def row_count(self) -> int:
        self._open()
        return self._row_count

    def get_row(self, idx: int) -> Dict[str, Any]:
        self._open()
        assert self._header is not None

    def get_row(self, idx: int) -> Dict[str, Any]:
        self._open()
        if self._data_cache is not None:
            if 0 <= idx < len(self._data_cache):
                return self._data_cache[idx]
            raise IndexError(f"Row {idx} out of range")
        
        # Fallback for .xls (xlrd)
        if self._is_xls:
            # For simplicity, we can also cache .xls if needed, but xlrd is already memory-resident
            # We just need to check if the user is using .xls and handle the valid_row_map.
            # However, modern users use .xlsx. Let's ensure basic support works.
            row = self._xlrd_ws.row_values(idx + 1)
            values = [str(v) if v is not None else "" for v in row]
            return dict(zip(self._header, values))

        raise IndexError(f"Row index {idx} not found")

    def iter_rows(self, start: int = 0, end: Optional[int] = None) -> Generator[Dict[str, Any], None, None]:
        self._open()
        if self._data_cache is not None:
            actual_end = end if end is not None else len(self._data_cache)
            for i in range(start, actual_end):
                yield self._data_cache[i]
            return

        # Fallback for .xls
        actual_end = end if end is not None else self._row_count
        if self._is_xls:
            for i in range(start, actual_end):
                row = self._xlrd_ws.row_values(i + 1)
                values = [str(v) if v is not None else "" for v in row]
                yield dict(zip(self._header, values))

    def close(self) -> None:
        if self._wb is not None:
            self._wb.close()
            self._wb = None
            self._ws = None

    def get_sheet_names(self) -> List[str]:
        """Return available sheet names in the workbook."""
        self._open()
        if self._is_xls:
            return self._xlrd_wb.sheet_names()
        return self._wb.sheetnames
